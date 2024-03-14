from math import trunc
import pathlib
import time
import sys
from datetime import datetime 
from datetime import date
import calendar
import openpyxl
from pathlib import Path
import sendWA
import logging


# Check post schedule day, time and process indicator 
#       to determine if the record should be processed
def is_time_to_send_msg(msgDay, msgTime):
    if(msgDay is None or msgTime is None):
        logger.error("ERROR: Either msgDay or  msgTime is None. Continuing to the next record..")
        return False

    hour,minute,second = msgTime.split(':')
    scheduleday = msgDay[:3].upper()
    sysday = datetime.today().strftime('%a').upper()
    currenttime = datetime.now()
    logger.debug(currenttime)
    if (sysday == scheduleday) or (scheduleday == "ALL"):
        if (currenttime.hour == int(hour)) and (currenttime.minute == int(minute)):
            logger.info("Message scheduled for today and for now. Starting the delivery..")
            return True
        else:
            logger.info("Message scheduled for today but it is not time yet.")
            return True
    else:
        logger.info("Message is not scheduled for delivery today")
        return True

def process_whatsapp_schedule():
    try:
        msgDict = {}
        #print("readExcelScheduler")
        xlsx_file = Path(Path().absolute(),'..\data\schedule_test.xlsx')

        wb_obj = openpyxl.load_workbook(xlsx_file)
        #sheet = wb_obj.active # load active sheet
        sheet = wb_obj["ScheduledMsgs"]
        col_names = []
        for column in sheet.iter_cols(1, sheet.max_column):
            col_names.append(column[0].value)
        #print(col_names)
        logger.info("Total records in the schedule: %s", sheet.max_row)
        #for row in sheet.iter_rows(max_row=sheet.max_row):
        sendMsgResult = None
        for r in range (2, sheet.max_row+1): #exclude first row (header)
            #print("row value",r)
            msgName = sheet.cell(r,1).value
            msgDay = sheet.cell(r,2).value
            msgTime = str(sheet.cell(r,3).value)
            msgText = sheet.cell(r,4).value
            logger.debug("msgText: %s", msgText)
            #Check if the msg is to be read from a text file
            if(msgText is not None and msgText.startswith('file:')):
                 msgText = read_msg_from_file(msgText[5:])
            #logger.info("printing message")
            #logger.info(msgText)
            msgPic = sheet.cell(r,5).value
            msgGroupAlias = sheet.cell(r,6).value
            msgDeliveryStatus = sheet.cell(r,7).value
            msgDict['msgName'] = msgName
            msgDict['msgText'] = msgText
            msgDict['msgPic'] = msgPic
            msgDict['msgGroupAlias'] = msgGroupAlias
            # print(msgDict)
            logger.info("Processing record:%s", r)
            logger.info("Message Name:%s, Scheduled Day:%s, Scheduled Time:%s", msgName, msgDay, msgTime)
            logger.info("msgDeliveryStatus:%s", msgDeliveryStatus)
            # check if it is time to send the msg
            #FIXME: below line commented for testing
            if((msgDay == "ALL" or msgDeliveryStatus == "SCHEDULED") and is_time_to_send_msg(msgDay, msgTime)):
            #if(True): # FIXME: comment this and uncomment above line
                # process msg
                sendMsgResult = sendWA.invoke_wahandler(msgDict)
                logger.info("sendMsgResult: %s", sendMsgResult)
            else:
                logger.error("Message does not qualify the scheduling criteria")
                continue
            if(sendMsgResult['msgStatus'] == "SUCCESSFUL"):
                logger.info("No errors sending message:%s", msgName )
                #update DeliveryStatus column with "SUCCESSFUL"
                sheet.cell(r,7).value = "SUCCESSFUL"
            elif(sendMsgResult['msgStatus'] == "FAILED"):
                logger.error("Error occurred while sending message")
                sheet.cell(r,7).value = "FAILED"
                sheet.cell(r,8).value = "Delivery failed to groups:" + sendMsgResult['failedToGrps']
            else:
                logger.error("Unknown return value sendMsgResult['msgStatus']:", sendMsgResult['msgStatus'])
        
            #save the file at the end of each record processing
            wb_obj.save(xlsx_file)
    
    except:
        logger.exception("Fatal error in process_whatsapp_schedule", exc_info=True)

def check_if_scheduler_running():
    print("check_if_scheduler_running")

def read_msg_from_file(fileName):
    logger.info("read_msg_from_file:%s",fileName)
    filePath = str(Path(dataPath,fileName))
    with open(filePath, 'r', encoding='utf8') as f:
        return f.read()

def setup_logging():
    logdir = "C:\\Logs"
    Path(logdir).mkdir(parents=True, exist_ok=True) #create log directory if not exists
    logfilename = "whatsappscheduler_"+datetime.today().strftime('%Y-%m-%d')+".log"
    logfileWithPath=str((Path(logdir,logfilename)))
    logging.basicConfig(level=logging.INFO, filename=logfileWithPath, format='%(asctime)s %(name)-12s %(levelname)-8s:%(message)s')

projectRoot = Path(__file__).parents[1]
dataPath = Path(projectRoot, "data")
logger = logging.getLogger(__name__)
setup_logging()
logger.info("********************** START PYTHON SCHEDULER **********************")
#start processing the scheduled whatsappp messages
process_whatsapp_schedule()
logger.info("********************** END PYTHON SCHEDULER ************************")
logging.shutdown()
exit(0)